from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msb
import xlrd
import openpyxl as xl
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import csv
import os
import webbrowser
import datetime
import statistics as stat

class PsaTsaProgram:
    def __init__(self):
        self.raw_data_location = '\\\\Ta1d181222\\01.result test peel strength\\DATA IPQC'
        self.master_list_location = "\\\\Ta1d180613\\2.MASTER LIST RUNNING NUMBER\\2. Master list IPQC"
        self.result_peel_location = "\\\\Ta1d180613\\3.RESULT DATA PSA\IPQC (SMT)\\2. Result peel strength (ผลการตรวจสอบจาก ANC)"
        self.spc_master_location = "\\\\Ta1d180613\\3.RESULT DATA PSA\IPQC (SMT)\\99.SPC Master(ห้ามลบ)\\QF-A1-QUA-2209-1.xlsx"
        self.tutorial_book = "\\\\Ta1d180613\\3.RESULT DATA PSA\\IPQC (SMT)\\100.โปรแกรม (ห้ามลบ)\\SMT PSA-TSA Hand Book.xlsx"
        self.control_folder_location = "\\\\Ta1d180613\\3.RESULT DATA PSA\\IPQC (SMT)\\3.Control limit Record"
        self.master_control_limit_record = "\\\\Ta1d180613\\3.RESULT DATA PSA\\IPQC (SMT)\\99.SPC Master(ห้ามลบ)\\Control limit Record.xlsx"

        # Widget Properties
        self.topic_font = ('Arial', 24, 'bold')
        self.detail_font = ('Arial', 16)

    def main_program(self):
        self.spc_window = Tk()
        WIDTH = 550
        HEIGHT = 580
        screen_width = self.spc_window.winfo_screenwidth()
        screen_height = self.spc_window.winfo_screenheight()
        x = int( (screen_width/2) - (WIDTH/2) )
        y = int( (screen_height/2) - (HEIGHT/1.8))

        # GUI Properties
        self.spc_window.title('IPQC-PSATSA Program')
        self.spc_window.resizable(0, 0)
        self.spc_window.geometry(f'{WIDTH}x{HEIGHT}+{x}+{y}')
        self.spc_window.iconbitmap('fujikura_logo.ico')

        # Menu bar
        menubar = Menu(self.spc_window)
        self.spc_window.config(menu=menubar)
        helpmenu = Menu(menubar,tearoff=0)
        menubar.add_cascade(label='Help',menu=helpmenu)
        helpmenu.add_command(label='โปรแกรม', command=lambda: msb.showinfo(title='แจ้งเตือนไปยังผู้ใช้', message='โปรแกรมสำหรับ SMT PSA-TSA Version 01'))
        helpmenu.add_command(label='วิธีการใช้',command=self.open_tutorial_book)
        helpmenu.add_separator()
        helpmenu.add_command(label='ติดต่อ',command=lambda: msb.showinfo('แจ้งเตือนไปยังผู้ใช้','หากโปรแกรมมีปัญหาหรือข้อสงสัยโทรเบอร์ 4308 ณัฐพล(แอ๊ค)'))

        # WIDGET Properties
        big_frame = Frame(self.spc_window)
        topic_label = Label(big_frame, text='SPC Report Program', font=self.topic_font, fg='white', bg='blue')
        select_frame = Frame(big_frame)
        year_label = Label(select_frame, text='Year Report:', font=self.detail_font)
        self.year_cb = ttk.Combobox(select_frame, justify='center', font=self.detail_font, width=18)
        month_label = Label(select_frame, text='Month Report:', font=self.detail_font)
        self.month_cb = ttk.Combobox(select_frame, justify='center', font=self.detail_font, width=18)
        product_label = Label(select_frame, text='Product Report:', font=self.detail_font)
        self.product_cb = ttk.Combobox(select_frame, justify='center', font=self.detail_font, width=18)
        machine_label = Label(select_frame, text='Machine No:', font=self.detail_font)
        machine_label = Label(select_frame, text='Machine No:', font=self.detail_font)
        self.machine_cb = ttk.Combobox(select_frame, justify='center', font=self.detail_font, width=18)
        doc_req_label = Label(select_frame, text='Report Name:', font=self.detail_font)
        check_box_frame = Frame(select_frame)
        self.peel_strength_name = StringVar()
        flex_check_box = Radiobutton(check_box_frame, text='Flex Report', font=self.detail_font, variable=self.peel_strength_name, value='flex')
        liner_check_box = Radiobutton(check_box_frame, text='Liner Report', font=self.detail_font, variable=self.peel_strength_name, value='liner')
        record_label = Label(select_frame, text='Record By:', font=self.detail_font)
        record_variable = StringVar()
        record_variable.set("SMT")
        button_style = ttk.Style()
        button_style.configure('my.TButton', font=('Arial', 18))
        self.record_entry = ttk.Entry(select_frame, font=self.detail_font, textvariable=record_variable, justify='center', width=19)
        run_program_button = ttk.Button(big_frame, text='Run Program', style='my.TButton', command=self.spc_program_overview)

        # Treeview Widget
        treeview_frame = Frame(big_frame)
        headers = ['Testing No.', 'Lot No.', 'Report status']
        self.psa_tas_treeview = ttk.Treeview(treeview_frame, column=headers, show='headings', height=7)
        vertical_scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=self.psa_tas_treeview.yview)
        self.psa_tas_treeview.configure(yscrollcommand=vertical_scrollbar.set)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 14, 'bold'), foreground="green")
        style.configure("Treeview", font=('Arial', 10))
        for header in headers:
            self.psa_tas_treeview.heading(header, text=header)
            self.psa_tas_treeview.column(header, anchor='center', width=160, minwidth=0)

        # Widget Event
        self.year_cb.bind('<Button-1>', self.year_click)
        self.month_cb.bind('<Button-1>', self.month_click)
        self.product_cb.bind('<Button-1>', self.product_click)
        self.machine_cb.bind('<Button-1>', self.machine_click)
        flex_check_box.invoke()

        # Widget Position
        big_frame.pack()
        topic_label.grid(row=0, column=0, pady=10, ipadx=10)
        select_frame.grid(row=1, column=0)
        year_label.grid(row=0, column=0, padx=5, pady=5)
        self.year_cb.grid(row=0, column=1, padx=5, pady=5, ipady=2)
        month_label.grid(row=1, column=0, padx=5, pady=5)
        self.month_cb.grid(row=1, column=1, padx=5, pady=5, ipady=2)
        product_label.grid(row=2, column=0, padx=5, pady=5)
        self.product_cb.grid(row=2, column=1, padx=5, pady=5, ipady=2)
        machine_label.grid(row=3, column=0, padx=5, pady=5)
        self.machine_cb.grid(row=3, column=1, padx=5, pady=5, ipady=2)
        doc_req_label.grid(row=4, column=0, padx=5, pady=5)
        check_box_frame.grid(row=4, column=1, padx=5, pady=5)
        flex_check_box.grid(row=0, column=0, padx=5)
        liner_check_box.grid(row=0, column=1, padx=5)
        record_label.grid(row=5, column=0, padx=5, pady=5)
        self.record_entry.grid(row=5, column=1, padx=5, pady=5, ipady=2)
        run_program_button.grid(row=2, column=0, pady=2, ipadx=20, ipady=3)
        treeview_frame.grid(row=4, column=0, pady=5)
        self.psa_tas_treeview.grid(row=0, column=0)
        vertical_scrollbar.grid(row=0, column=1, ipady=60)

        # GUI Activate
        self.spc_window.mainloop()

    def open_tutorial_book(self):
        try:
            webbrowser.open(url=self.tutorial_book)
        except:
            pass

    def year_click(self, event):
        folder_name_list = []

        for folder_name in os.listdir(self.raw_data_location):
            if str(folder_name).startswith("20"):
                folder_name_list.append(folder_name)

        self.year_cb['values'] = folder_name_list
        self.month_cb.set("")
        self.product_cb.set("")
        self.machine_cb.set("")
    
    def month_click(self, event):
        year_input = self.year_cb.get()

        if year_input != "":
            month_all_folders_location = os.path.join(self.raw_data_location, year_input)
            month_folders_list = os.listdir(month_all_folders_location)

            self.month_cb['values'] = month_folders_list
            self.product_cb.set("")
            self.machine_cb.set("")
        else:
            msb.showwarning('แจ้งเตือนไปยังผู้ใช้', 'กรุณากรอกข้อมูลที่ช่อง Year Report ก่อน')
            self.month_cb['values'] = []
    
    def product_click(self, event):
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()

        if year_input != "" and month_input != "": 
            month_input_list = month_input.split("'")
            month_name = month_input_list[1][0:3]
            new_master_file_location = os.path.join(self.master_list_location, year_input)
            product_name_list = []

            for excel_file_name in os.listdir(new_master_file_location):
                if month_name in excel_file_name and not excel_file_name.startswith("~$") and excel_file_name.endswith(".xlsx"):
                    master_file_location = os.path.join(new_master_file_location, excel_file_name)
                    # Open master file
                    master_book = xlrd.open_workbook(filename=master_file_location)
                    master_sheet = master_book.sheet_by_index(0)
                    start_row = 8
                    product_name_cell = master_sheet.cell(rowx=start_row, colx=3).value

                    while product_name_cell != "":
                        product_name = str(product_name_cell)[0:10]

                        # Add data to list when product startwith RG and don't have add to list
                        if product_name.upper() not in product_name_list and product_name.upper().startswith('RG'):
                            product_name_list.append(product_name.upper())

                        start_row += 5
                        product_name_cell = master_sheet.cell(rowx=start_row, colx=3).value

                    # Close Master file
                    master_book.release_resources()
            
            self.product_cb['values'] = product_name_list
            self.machine_cb.set("")
        else:
            msb.showwarning('แจ้งเตือนไปยังผู้ใช้', 'กรุณากรอกข้อมูลที่ช่อง Year Report และ Month Report ก่อน')
            self.product_cb['values'] = []
            self.product_cb.set("")

    def machine_click(self, event):
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()

        if year_input != "" and month_input != "" and product_input != "": 
            month_input_list = month_input.split("'")
            month_name = month_input_list[1][0:3]
            new_master_file_location = os.path.join(self.master_list_location, year_input)
            machine_number_list = []
            
            for excel_file_name in os.listdir(new_master_file_location):
                if month_name in excel_file_name and not excel_file_name.startswith("~$") and excel_file_name.endswith(".xlsx"):
                    master_file_location = os.path.join(new_master_file_location, excel_file_name)
                    # Open master file
                    master_book = xlrd.open_workbook(filename=master_file_location)
                    master_sheet = master_book.sheet_by_index(0)
                    start_row = 8
                    product_name_cell = master_sheet.cell(rowx=start_row, colx=3).value
                    machine_number_cell = master_sheet.cell(rowx=start_row, colx=5).value

                    while product_name_cell != "":
                        product_name = str(product_name_cell).upper()
                        
                        if product_name.startswith(product_input) and machine_number_cell.strip() not in machine_number_list:
                            machine_number_list.append(str(machine_number_cell).strip())

                        # Loop command
                        start_row += 5
                        product_name_cell = master_sheet.cell(rowx=start_row, colx=3).value
                        machine_number_cell = master_sheet.cell(rowx=start_row, colx=5).value

                    # Close master file
                    master_book.release_resources()

            self.machine_cb['values'] = machine_number_list
        else:
            msb.showwarning('แจ้งเตือนไปยังผู้ใช้', 'กรุณากรอกข้อมูลที่ช่อง Year Report Month Report และ Product Report ก่อน')
            self.machine_cb['values'] = []
            self.machine_cb.set("")

    def spc_program_overview(self):
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
    
        self.clear_psa_tsa_treeview()
        
        if year_input == "" and month_input == "" and product_input == "" and machine_input == "":
            msb.showinfo(title='แจ้งเตือนไปยังผู้ใช้', message='คุณกรอกข้อมูลข้างบนไม่ครบถ้วน')
        else:
            # Function Overview
            master_list = self.search_and_open_master_file()
            
            if len(master_list) != 0:
                spc_file_location, result_peel_location2 = self.search_spc_file_location()
                spc_book, spc_sheet1, spc_sheet2, filename_open= self.open_spc_file(spc_file_location)
                self.check_iqic_has_update(spc_book, spc_sheet1, spc_sheet2, master_list)
                self.close_and_save_spc_file(spc_book, spc_sheet1, spc_file_location, result_peel_location2, filename_open, master_list)
                
            else:
                msb.showinfo(title='แจ้งเตือนไปยังผู้ใช้', message='ยังไม่มีข้อมูลสำหรับการอัพเดท')
            
    def search_and_open_master_file(self):
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        month_input = self.month_cb.get()
        year_input = self.year_cb.get()
        peel_strength_input = self.peel_strength_name.get()

        month_input_list = month_input.split("'")
        month_name = month_input_list[1][0:3]
        new_master_file_location = os.path.join(self.master_list_location, year_input)

        for excel_file_name in os.listdir(new_master_file_location):
            if month_name in excel_file_name and not excel_file_name.startswith("~$") and excel_file_name.endswith(".xlsx"):
                # Open Excel file
                master_file_location = os.path.join(new_master_file_location, excel_file_name)
                master_book = xlrd.open_workbook(filename=master_file_location)
                master_sheet = master_book.sheet_by_index(0)

                # Run function
                master_list = self.get_information_in_master(master_book, master_sheet)
                
                master_book.release_resources()
                break

        return master_list

    def get_information_in_master(self, master_book, master_sheet):
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        peel_strength_input = self.peel_strength_name.get()
        master_list = []

        start_row = 8
        end_row = master_sheet.nrows
        product_name = master_sheet.cell(rowx=start_row, colx=3).value

        while str(product_name) != "":
            testing_no = str(master_sheet.cell(rowx=start_row, colx=1).value).strip()
            request_date_cell = master_sheet.cell(rowx=start_row, colx=2).value
            request_date = datetime.datetime(*xlrd.xldate_as_tuple(request_date_cell, master_book.datemode)).strftime('%x')
            lot_no = int(master_sheet.cell(rowx=start_row, colx=4).value)
            auto_press_machine = master_sheet.cell(rowx=start_row, colx=5).value

            if machine_input in auto_press_machine and product_input in product_name:
                # Get Serial Bar Code list
                serial_barcode_list = []
                if peel_strength_input == 'flex':
                    peel_column = 8
                else:
                    peel_column = 10
                for peel_row in range(start_row, start_row+5):
                    serial_barcode = master_sheet.cell(rowx=peel_row, colx=peel_column).value
                    serial_barcode_list.append(serial_barcode)

                # Run function
                peel_result_list = self.search_raw_file_location(testing_no, lot_no)
                if len(peel_result_list) == 5:
                    data_list = [testing_no, request_date, product_name, lot_no, auto_press_machine, serial_barcode_list, peel_result_list]
                    master_list.append(data_list)

            # Loop command
            start_row += 5
            product_name = master_sheet.cell(rowx=start_row, colx=3).value
        
        return master_list

    def search_raw_file_location(self, testing_no, lot_no):
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        peel_strength_input = self.peel_strength_name.get()
        product_input = self.product_cb.get()
        peel_result_list = []
        
        iqic_folder_location = self.raw_data_location + "\\" + year_input + "\\" + month_input + "\\" + testing_no
        if os.path.isdir(iqic_folder_location):
            for file in os.listdir(iqic_folder_location):
                # Check case file name
                if "SP" in file.upper():
                    continue
                elif peel_strength_input == "flex" and file.upper().endswith("CSV") and "F" in file:
                    csv_file_location = iqic_folder_location + "\\" + file
                    peel_result_list = self.get_list_when_open_csv(csv_file_location)
                    break
                elif peel_strength_input == "flex" and file.upper().endswith("XLS") and "F" in file:
                    excel_file_location = iqic_folder_location + "\\" + file
                    peel_result_list = self.get_list_when_open_excel(excel_file_location)
                    break
                elif peel_strength_input == "liner" and file.upper().endswith("CSV") and "L" in file:
                    csv_file_location = iqic_folder_location + "\\" + file
                    peel_result_list = self.get_list_when_open_csv(csv_file_location)
                    break
                elif peel_strength_input == "liner" and file.upper().endswith("XLS") and "L" in file:
                    excel_file_location = iqic_folder_location + "\\" + file
                    peel_result_list = self.get_list_when_open_excel(excel_file_location)
                    break

        return peel_result_list

    def get_list_when_open_excel(self, excel_file_location):
        peel_book = xlrd.open_workbook(filename=excel_file_location)
        peel_sheet = peel_book.sheet_by_index(0)
        start_row = 0
        end_row = peel_sheet.nrows - 1
        peel_result_list = []

        while start_row <= end_row:
            title_name = peel_sheet.cell(rowx=start_row, colx=0).value

            if isinstance(title_name, float):
                peel_result = float(peel_sheet.cell(rowx=start_row, colx=1).value)
                peel_result_list.append(peel_result)

            start_row += 1

        peel_book.release_resources()
        
        return peel_result_list

    def get_list_when_open_csv(self, csv_file_location):
        item_no_list = ['1', '2', '3', '4', '5']
        peel_result_list = []

        with open(csv_file_location) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')

            for row in csv_reader:
                item_no = row[0]
                if item_no in item_no_list:
                    peel_result = float(row[1])
                    peel_result_list.append(peel_result)
                    if item_no == '5':
                        break

        csv_file.close()
        
        return peel_result_list

    def search_spc_file_location(self):
        spc_file_location = ""
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        peel_strength_input = self.peel_strength_name.get()

        # Transform to folder name and path name
        year_folder_name = 'Data' + year_input
        result_peel_location1 = os.path.join(self.result_peel_location, year_folder_name)
        month_folder_list = month_input.split("'")
        month_folder_name = month_folder_list[0] + "." + month_folder_list[1]
        result_peel_location2 = os.path.join(result_peel_location1, month_folder_name)

        for spc_file in os.listdir(result_peel_location2):
            if not spc_file.startswith('~$') and spc_file.startswith('FLEX') and peel_strength_input == 'flex' and product_input in spc_file and machine_input in spc_file:
                spc_file_location = os.path.join(result_peel_location2, spc_file)
            elif not spc_file.startswith('~$') and spc_file.startswith('LINER') and peel_strength_input == 'liner' and product_input in spc_file and machine_input in spc_file:
                spc_file_location = os.path.join(result_peel_location2, spc_file)

        return spc_file_location, result_peel_location2

    def open_spc_file(self, spc_file_location):
        if spc_file_location == "":
            filename_open= self.spc_master_location
        else:
            filename_open = spc_file_location

        spc_book = xl.open(filename=filename_open)
        spc_sheet_list = spc_book.sheetnames
        spc_sheet1 = spc_book[spc_sheet_list[0]]
        spc_sheet2 = spc_book[spc_sheet_list[1]]

        return spc_book, spc_sheet1, spc_sheet2, filename_open

    def clear_psa_tsa_treeview(self):
        for member in self.psa_tas_treeview.get_children():
            self.psa_tas_treeview.delete(member)

    def check_iqic_has_update(self, spc_book, spc_sheet1, spc_sheet2, master_list):
        col_spc_sheet2 = 2

        for row in master_list:
            iqic_no = row[0]
            lot_no = row[3]

            has_upate_at_spc_sheet2 = self.iqic_no_has_update_in_spc_sheet2(spc_sheet2 , master_list, lot_no)
            
            if has_upate_at_spc_sheet2:
                data = [iqic_no, lot_no, "อัพเดทไปแล้ว"]
            else:
                data = [iqic_no, lot_no, "โปรแกรมอัพเดท"]
                self.update_result_to_spc_sheet1(spc_sheet1, row)
                self.update_result_to_spc_sheet2(spc_sheet2, row)
            
            self.psa_tas_treeview.insert('', 'end', value=data)

    def iqic_no_has_update_in_spc_sheet2(self, spc_sheet2 , master_list, lot_no):
        has_upate_at_spc_sheet2 = False
        spc_column = 1
        lot_no_in_spc_sheet2 = spc_sheet2.cell(row=4, column=spc_column).value
        
        while str(lot_no_in_spc_sheet2) != "None":
            if str(lot_no) in str(lot_no_in_spc_sheet2):
                has_upate_at_spc_sheet2 = True
                return has_upate_at_spc_sheet2

            # Loop command
            spc_column += 1
            lot_no_in_spc_sheet2 = spc_sheet2.cell(row=4, column=spc_column).value
            column_update_in_spc_sheet1 = spc_column

        return has_upate_at_spc_sheet2

    def update_result_to_spc_sheet1(self, spc_sheet1, master_row):
        # Search Last Column Update
        spc_column = 3
        peel_result = spc_sheet1.cell(row=55, column=spc_column).value

        while str(peel_result) != "None":
            # Loop command
            spc_column += 1
            peel_result = spc_sheet1.cell(row=55, column=spc_column).value
            
        # Update data 
        peel_result_list = master_row[6]
        spc_row = 55

        for peel_result in peel_result_list:
            spc_sheet1.cell(row=spc_row, column=spc_column).value = float(peel_result)
            spc_row += 1

    def update_result_to_spc_sheet2(self, spc_sheet2, master_row):
        # Search Last Column Update
        spc_column = 2
        lot_no = spc_sheet2.cell(row=4, column=spc_column).value

        while str(lot_no) != "None":
            # Loop command
            spc_column += 1
            lot_no = spc_sheet2.cell(row=4, column=spc_column).value
        
        # Update data request_date, lot_no and serial barcode
        request_date = master_row[1]
        lot_no = master_row[3]
        serial_barcode_list = master_row[5]
        spc_row = 5

        spc_sheet2.cell(row=3, column=spc_column).value = request_date
        spc_sheet2.cell(row=4, column=spc_column).value = lot_no
        for serial_barcode in serial_barcode_list:
            spc_sheet2.cell(row=spc_row, column=spc_column).value = serial_barcode
            spc_sheet2.cell(row=13, column=spc_column).value = self.record_entry.get()
            spc_row += 1

    def import_spec_control_limit(self, last_spc_file_location, spc_sheet1):
        last_spc_book = xlrd.open_workbook(filename=last_spc_file_location)
        last_spc_sheet = last_spc_book.sheet_by_index(0)

    def close_and_save_spc_file(self, spc_book, spc_sheet1, spc_file_location, result_peel_location2, filename_open, master_list):
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        peel_strength_input = self.peel_strength_name.get()
        new_date_format = month_input.split("'")[1] + "'" + year_input[2:4]
        
        # Set Default location and name to save file
        if spc_file_location == "" and peel_strength_input == 'flex':
            new_save_file_name = 'FLEX PEEL STRENGTH_' + product_input + "_" + machine_input + "_" + new_date_format + ".xlsx"
            new_save_file_location = os.path.join(result_peel_location2, new_save_file_name)
            self.update_spc_information_detail(spc_sheet1, new_date_format, machine_input, product_input, new_save_file_location)
            self.record_last_control_limit_to_spc(spc_sheet1)

        elif spc_file_location == "" and peel_strength_input == 'liner':
            new_save_file_name = 'LINER PEEL STRENGTH_' + product_input + "_" + machine_input + "_" + new_date_format + ".xlsx"
            new_save_file_location = os.path.join(result_peel_location2, new_save_file_name)
            self.update_spc_information_detail(spc_sheet1, new_date_format, machine_input, product_input, new_save_file_location)
            self.record_last_control_limit_to_spc(spc_sheet1)

        else:
            new_save_file_location = spc_file_location


        # Save file and Close Workbook
        try:
            spc_book.save(new_save_file_location)
            spc_book.close()

            self.open_control_limit_record(master_list)

            # Run function for open SPC file
            self.ask_open_file_name(new_save_file_location)
        except Exception as e:
            msb.showwarning(title='แจ้งเตือนไปยังผู้ใช้', message=f'กรุณาปิดไฟล์ \n {os.path.basename(new_save_file_location)} \n แล้วทำการรันโปรแกรมใหม่อีกครั้ง')

    def update_spc_information_detail(self, spc_sheet1, new_date_format, machine_input, product_input, new_save_file_location):
        spc_sheet1.cell(row=2, column=3).value = 'IPQC PSA, TSA'
        spc_sheet1.cell(row=3, column=3).value = 'FLEX PEELSTRENGTH'
        spc_sheet1.cell(row=4, column=3).value = machine_input
        spc_sheet1.cell(row=5, column=3).value = '-'

        spc_sheet1.cell(row=2, column=7).value = product_input
        spc_sheet1.cell(row=3, column=7).value = new_date_format
        spc_sheet1.cell(row=4, column=7).value = '-'
        spc_sheet1.cell(row=5, column=7).value = "5 PCS / SHIFT"

    def record_last_control_limit_to_spc(self, spc_sheet1):
        year_input = self.year_cb.get()
        last_year_input = int(year_input) - 1
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        report_input = self.peel_strength_name.get()

        if "JAN" in month_input:
            control_limit_folder_path = self.control_folder_location + "\\" + str(last_year_input)
            control_limit_file_name = "Control limit Record " + str(last_year_input) + ".xlsx"
        else:
            control_limit_folder_path = self.control_folder_location + "\\" + year_input
            control_limit_file_name = "Control limit Record " + year_input + ".xlsx"

        control_limit_file_location = os.path.join(control_limit_folder_path, control_limit_file_name)
        if os.path.isfile(control_limit_file_location):
            control_book = xlrd.open_workbook(filename=control_limit_file_location)
            control_sheet = control_book.sheet_by_index(0)
            min_row = 0
            max_row = control_sheet.nrows - 1

            # Search Row Record
            while max_row >= min_row:
                month_in_table = control_sheet.cell(rowx=max_row, colx=1).value
                product_name_in_table = control_sheet.cell(rowx=max_row, colx=2).value
                machine_in_table = control_sheet.cell(rowx=max_row, colx=3).value
                report_name_in_table = control_sheet.cell(rowx=max_row, colx=4).value

                condition_filled = (product_input == product_name_in_table and machine_input == machine_in_table and report_input == report_name_in_table and month_input != month_in_table)

                if condition_filled:
                    # X chart Control Limit
                    uclx = control_sheet.cell(rowx=max_row, colx=5).value
                    clx = control_sheet.cell(rowx=max_row, colx=6).value
                    lclx = control_sheet.cell(rowx=max_row, colx=7).value
                    # R chart Control Limit
                    uclr = control_sheet.cell(rowx=max_row, colx=8).value
                    clr = control_sheet.cell(rowx=max_row, colx=9).value
                    lclr = control_sheet.cell(rowx=max_row, colx=10).value

                    # Filled Last control limit to spc
                    spc_sheet1.cell(row=4, column=12).value = clx
                    spc_sheet1.cell(row=5, column=12).value = uclx
                    spc_sheet1.cell(row=6, column=12).value = lclx
                    spc_sheet1.cell(row=4, column=14).value = clr
                    spc_sheet1.cell(row=5, column=14).value = uclr
                    spc_sheet1.cell(row=6, column=14).value = lclr
                    break

                # Loop command
                max_row -= 1
            
            control_book.release_resources()

    def open_control_limit_record(self, master_list):
        year_input = self.year_cb.get()

        control_limit_location = self.control_folder_location + "\\" + year_input + "\\Control limit Record " + year_input + ".xlsx"
        if os.path.isfile(control_limit_location):
            filename_record = control_limit_location
        else:
            filename_record = self.master_control_limit_record

        try:
            control_book = xl.load_workbook(filename=filename_record)
            control_sheet = control_book[control_book.sheetnames[0]]

            # Run function
            uclx, clx, lclx, uclr, clr, lclr = self.calculate_current_control_limit(master_list)
            self.record_control_limit(uclx, clx, lclx, uclr, clr, lclr, control_sheet)

            control_book.save(filename=filename_record)
            control_book.close()
        except Exception:
            msb.showwarning('แจ้งเตือนไปยังผู้ใช้', message=f'กรุณาปิดไฟล์ที่ชื่อ \n {os.path.basename(filename_record)} \n แล้วทำการรันใหม่อีกครั้ง')

    def calculate_current_control_limit(self, master_list):
        xbar_r = []
        r = []

        # Calculate Xbar i and R i of each subgroup
        for row in master_list:
            peel_result_list = row[6]
            avg = stat.mean(peel_result_list)
            # rng = range = max - min
            rng = max(peel_result_list) - min(peel_result_list)
            xbar_r.append(avg)
            r.append(rng)

        xbarbar = stat.mean(xbar_r)
        rbar = stat.mean(r)

        # Calculate Xbar formula
        uclx = round(xbarbar + (0.577 * rbar), 3)
        clx = round(xbarbar, 3)
        lclx = round(xbarbar - (0.577 * rbar), 3)
        # Calculate R formula
        uclr = round(2.114 * rbar, 3)
        clr = round(rbar, 3)
        lclr = 0

        return uclx, clx, lclx, uclr, clr, lclr

    def record_control_limit(self, uclx, clx, lclx, uclr, clr, lclr, control_sheet):
        has_filled = False
        year_input = self.year_cb.get()
        month_input = self.month_cb.get()
        product_input = self.product_cb.get()
        machine_input = self.machine_cb.get()
        report_input = self.peel_strength_name.get()
        record_by_input = self.record_entry.get()

        max_row = control_sheet.max_row
        min_row = 1
        while min_row <= max_row:
            month_table = control_sheet.cell(row=min_row, column=2).value
            product_table = control_sheet.cell(row=min_row, column=3).value
            machine_table = control_sheet.cell(row=min_row, column=4).value
            report_name_table = control_sheet.cell(row=min_row, column=5).value

            update_condition = (month_input == month_table and product_input == product_table and machine_input == machine_table and report_name_table == report_input)
            if update_condition:
                has_filled = True
                control_sheet.cell(row=min_row, column=6).value = uclx
                control_sheet.cell(row=min_row, column=7).value = clx
                control_sheet.cell(row=min_row, column=8).value = lclx
                control_sheet.cell(row=min_row, column=9).value = uclr
                control_sheet.cell(row=min_row, column=10).value = clr
                control_sheet.cell(row=min_row, column=11).value = lclr
                break

            # Loop command
            min_row += 1

        if has_filled == False:
            last_filled_row = int(control_sheet.max_row) + 1
            list_filled = [year_input, month_input, product_input, machine_input, report_input, 
                          uclx, clx, lclx, uclr, clr, lclr, record_by_input]

            filled_col = 1
            for item in list_filled:
                control_sheet.cell(row=last_filled_row, column=filled_col).value = item

                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center')
                control_sheet.cell(row=last_filled_row, column=filled_col).alignment = center_align
                control_sheet.cell(row=last_filled_row, column=filled_col).border = thin_border

                filled_col += 1

    def ask_open_file_name(self, new_save_file_location):
        need_open_spc_file = msb.askyesno(title='แจ้งเตือนไปยังผู้ใช้', message='คุณต้องการจะเปิดไฟล์ \n' + os.path.basename(new_save_file_location))

        if need_open_spc_file:
            webbrowser.open(url=new_save_file_location)


app = PsaTsaProgram()
app.main_program()