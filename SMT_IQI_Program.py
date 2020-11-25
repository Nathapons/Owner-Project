from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msb
import openpyxl as xl
import webbrowser
import csv
import random
import xlrd
import os
import shutil


class SmtIqiProgram:
    def __init__(self):
        # Adhesive link path
        self.adhesive_machine_path = "\\\\ta1d181222\\D\\01.RESULT TEST PEEL STRENGTH\\DATA IQI"
        adhesive_format_path1 = "\\\\ta1d170805\\02.TEST ADH PAD (IQI-SMT)ห้ามลบ"
        adhesive_format_path2 = "01. Thickness & Peel strength Result iqi-smt"
        self.adhesive_format_path = os.path.join(adhesive_format_path1, adhesive_format_path2)
        # Pull_force link path
        self.pull_force_machine_path = "\\\\ta1d181222\\D\\01.RESULT TEST PEEL STRENGTH\\DATA IQI\\CONNECTOR"
        # Copanarity link path
        self.copanarity_machine_path = "\\\\ta1d181222\\02.RESULT TEST 3D\\02. DATA_3D IQI SMT"
        self.copanarity_format_path = "\\\\ta1d170805\\01.Test ok 2 buid  IQI-SMT\\01.TEST OK2 BUILD (IQI-SMT)"

        # Text size of GUI Properties
        self.topic_font = ('Arial', 28, 'bold')
        self.creator_font = ('Arial', 16)
        self.detail_font = ('Arial', 18)
        self.combobox_width = 18
        self.treeview_height = 11

    def main_window(self):
        self.root = Tk()
        self.root.title('Main Window Program')
        WIDTH = 400
        HEIGHT = 230
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = int((screen_width/2) - (WIDTH/2))
        y = int((screen_height/2) - (HEIGHT/1.6))
        self.root.geometry(f'{WIDTH}x{HEIGHT}+{x}+{y}')
        self.root.resizable(0, 0)

        # Widget in GUI
        big_frame = Frame(self.root)
        topics_label = Label(big_frame, text='Please Select Menu', font=self.topic_font, fg='white', bg='blue')
        creator_label = Label(big_frame, text='Create by Nathapon.S Tel. 4955', font=self.creator_font)
        # adhesive_button = Button(big_frame, text='Adhesive Program', command="self.adhesive_window",
                                #  font=self.detail_font, width=20)
        pull_force_button = Button(big_frame, text='Pull Force Program', command=self.pull_force_window,
                                   font=self.detail_font, width=20)
        coparnarity_button = Button(big_frame, text='Coparnarity Program', command=self.copanarity_window,
                                    font=self.detail_font, width=20)

        # Widget place
        big_frame.pack()
        topics_label.grid(row=0, column=0, ipadx=20, pady=5)
        creator_label.grid(row=1, column=0, pady=5)
        #adhesive_button.grid(row=2, column=0, pady=5)
        pull_force_button.grid(row=3, column=0, pady=5)
        coparnarity_button.grid(row=4, column=0, pady=5)
        # Activate GUI
        self.root.mainloop()

    def hide_main_window(self):
        self.root.withdraw()

    # ---------------------------------------- Adhesive Program ----------------------------------------
    def adhesive_window(self):
        self.hide_main_window()

        self.adhesive_gui = Toplevel()
        self.adhesive_gui.title('Adhesive Program')
        WIDTH = 700
        HEIGHT = 500
        screen_width = self.adhesive_gui.winfo_screenwidth()
        screen_height = self.adhesive_gui.winfo_screenheight()
        x = int((screen_width / 2) - (WIDTH / 2))
        y = int((screen_height / 2) - (HEIGHT / 1.8))
        self.adhesive_gui.resizable(0, 0)
        self.adhesive_gui.geometry(f'{WIDTH}x{HEIGHT}+{x}+{y}')
        self.adhesive_gui.protocol("WM_DELETE_WINDOW", self.adhesive_close_window)
        # Widget
        big_frame = Frame(self.adhesive_gui)
        topics_label = Label(big_frame, text='Adhesive Program', font=self.topic_font, fg='white', bg='blue')
        # Widget Position
        big_frame.pack()
        topics_label.grid(row=0, column=0, pady=5)

    def adhesive_close_window(self):
        self.adhesive_gui.destroy()
        self.main_window()

    # --------------------------------------- Pull Force Program ----------------------------------------
    def pull_force_window(self):
        self.hide_main_window()
        self.pull_force_gui = Toplevel()
        WIDTH = 900
        HEIGHT = 550
        screen_width = self.pull_force_gui.winfo_screenwidth()
        screen_height = self.pull_force_gui.winfo_screenheight()
        x = int((screen_width/2) - (WIDTH/2))
        y = int((screen_height/2) - (HEIGHT/1.8))
        self.pull_force_gui.title('Pull Force Program')
        self.pull_force_gui.geometry(f'{WIDTH}x{HEIGHT}+{x}+{y}')
        self.pull_force_gui.protocol("WM_DELETE_WINDOW", self.close_pull_force_window)
        self.pull_force_gui.resizable(0, 0)

        # Widget
        big_frame = Frame(self.pull_force_gui)
        topics_label = Label(big_frame, text='Pull Force Program', font=self.topic_font, fg='white', bg='blue')
        filter_frame = Frame(big_frame)
        year_label = Label(filter_frame, text='Year Report', font=self.detail_font)
        self.year_cb = ttk.Combobox(filter_frame, font=self.detail_font, value=os.listdir(self.pull_force_machine_path),
                                    justify='center', width=self.combobox_width)
        month_label = Label(filter_frame, text='Month Report', font=self.detail_font)
        self.month_cb = ttk.Combobox(filter_frame, font=self.detail_font, justify='center', width=self.combobox_width)
        iqid_no_label = Label(filter_frame, text='IQID no.', font=self.detail_font)
        self.iqid_no_cb = ttk.Combobox(filter_frame, font=self.detail_font, justify='center', width=self.combobox_width)
        run_button = Button(big_frame, text='Run Program', command=self.pull_force_overview,
                            font=self.detail_font, width=self.combobox_width)
        treeview_frame = Frame(big_frame)
        headers = ['Filename', 'Item Code', 'FETL lot', 'Report status']
        self.pull_force_treeview = ttk.Treeview(treeview_frame, column=headers, show='headings',
                                                height=self.treeview_height)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 16, 'bold'))
        style.configure("Treeview", font=('Arial', 12))
        vertical_scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=self.pull_force_treeview.yview)
        self.pull_force_treeview.configure(yscrollcommand=vertical_scrollbar.set)
        for header in headers:
            self.pull_force_treeview.heading(header, text=header)
            if header == 'Filename':
                column_width = 360
            else:
                column_width = 160
            self.pull_force_treeview.column(header, anchor='center', width=column_width, minwidth=0)
            self.pull_force_treeview.bind("<Double-1>", self.pull_force_link_tree)

        # Function link to widget
        self.month_cb.bind("<Button-1>", self.pull_force_month_function)
        self.iqid_no_cb.bind("<Button-1>", self.pull_force_iqid_function)
        run_button.bind("<Return>", self.pull_force_overview)

        # Widget Position
        big_frame.pack()
        topics_label.grid(row=0, column=0, pady=10)
        # -- Filter frame
        filter_frame.grid(row=1, column=0)
        year_label.grid(row=0, column=0, pady=5, padx=5)
        self.year_cb.grid(row=0, column=1, pady=5, padx=5, ipady=2)
        month_label.grid(row=1, column=0, pady=5, padx=5)
        self.month_cb.grid(row=1, column=1, pady=5, padx=5, ipady=2)
        iqid_no_label.grid(row=2, column=0, pady=5, padx=5)
        self.iqid_no_cb.grid(row=2, column=1, pady=5, padx=5, ipady=2)
        run_button.grid(row=2, column=0, pady=5)
        treeview_frame.grid(row=3, column=0, pady=2)
        self.pull_force_treeview.grid(row=0, column=0)
        vertical_scrollbar.grid(row=0, column=1, ipady=100)

    def close_pull_force_window(self):
        self.pull_force_gui.destroy()
        self.main_window()

    def pull_force_month_function(self, event):
        year = self.year_cb.get()
        if year == "":
            msb.showwarning(title='แจ้งเตือน', message='กรุณากรอกช่อง Year Report ให้เรียบร้อย')
        else:
            link = os.path.join(self.pull_force_machine_path, year)
            self.month_cb['values'] = os.listdir(link)

    def pull_force_iqid_function(self, event):
        year = self.year_cb.get()
        month = self.month_cb.get()
        if year == "" and month == "":
             msb.showwarning(title='แจ้งเตือน', message='กรุณากรอกช่อง Year Report และ Month Report ให้เรียบร้อย')
        else:
            link = os.path.join(os.path.join(self.pull_force_machine_path, year), self.month_cb.get())
            self.iqid_no_cb['values'] = os.listdir(link)

    def pull_force_link_tree(self, event):
        input_id = self.pull_force_treeview.selection()
        item_code = str(self.pull_force_treeview.item(input_id)["values"][1])
        fetl_lot = self.pull_force_treeview.item(input_id)["values"][2]
        try:
            report_status, pull_force_format_full_path = self.search_format_file(item_code, fetl_lot)
            webbrowser.open(pull_force_format_full_path)
        except Exception:
            msb.showwarning(title="Alarm to message", messsage=f'FETL Lot {fetl_lot} ไม่สามารถเปิดได้เนื่องจากไม่ได้สร้างไฟล์เอาไว้')

    def pull_force_overview(self):
        total_user_not_fill, status = self.pull_count_user_not_fill()
        if total_user_not_fill != 0:
            msb.showwarning(title='แจ้งเตือน', message='กรุณากรอกข้อมูลที่ ' + status + ' ก่อนรันโปรแกรม')
        else:
            self.clear_pull_force_tree()
            self.update_data_to_pull_force_tree()
            msb.showinfo(title='Information', message='Import complete')

    def pull_count_user_not_fill(self):
        user_fill_dict = {"Year Report": self.year_cb.get(),
                          "Month Report": self.month_cb.get(),
                          "IQID no.": self.iqid_no_cb.get()}
        total_user_not_fill = 0
        status = " "
        for user_fill in user_fill_dict:
            if user_fill_dict[user_fill] == "":
                status = status + " " + user_fill
                total_user_not_fill += 1
        return total_user_not_fill, status

    def clear_pull_force_tree(self):
        for member in self.pull_force_treeview.get_children():
            self.pull_force_treeview.delete(member)

    def update_data_to_pull_force_tree(self):
        pull_force_path = (self.pull_force_machine_path + "\\"
                           + self.year_cb.get() + "\\"
                           + self.month_cb.get() + "\\"
                           + self.iqid_no_cb.get())
        for file_name in os.listdir(pull_force_path):
            if str(file_name).endswith('csv'):
                csv_file = os.path.join(pull_force_path, file_name)
                item_code, fetl_lot = self.get_fetl_and_itemcode(file_name)
                report_status, pull_force_format_full_path = self.search_format_file(item_code, fetl_lot)

                if report_status == "Update" and "~$" not in str(file_name):
                    # Run function
                    format_book, format_sheet = self.open_pull_force_format(pull_force_format_full_path)
                    format_row, format_column = self.pull_force_get_row_column(format_sheet)
                    report_status = self.open_close_pull_force_result(csv_file, format_sheet, format_row, format_column)
                    self.close_and_save_file(format_book, pull_force_format_full_path)

                    # Update data to Treeview
                    data = [file_name, item_code, fetl_lot, report_status]
                    self.pull_force_treeview.insert('', 'end', value=data)

    def get_fetl_and_itemcode(self, csv_file):
        first_char_in_itemcode = str(csv_file).index("-") + 1
        first_char_in_fetl_lot = str(csv_file).index("-", first_char_in_itemcode, len(csv_file)-1) + 1
        item_code = csv_file[first_char_in_itemcode: first_char_in_itemcode+8]
        fetl_lot = csv_file[first_char_in_fetl_lot: first_char_in_fetl_lot+8]
        return item_code, fetl_lot

    def open_pull_force_format(self, pull_force_format_full_path):
        format_book = xl.load_workbook(filename=pull_force_format_full_path)
        format_sheet = format_book[format_book.sheetnames[0]]

        return format_book, format_sheet

    def pull_force_get_row_column(self, format_sheet):
        for row in range(1, format_sheet.max_row):
            for column in range(1, format_sheet.max_column):
                cell_value = str(format_sheet.cell(row=row, column=column).value)
                if cell_value == "1":
                    format_row = row
                elif cell_value == "Pull Force":
                    format_column = column
        return format_row, format_column

    def open_close_pull_force_result(self, csv_file, format_sheet, format_row, format_column):
        with open(csv_file) as pull_force_result:
            pull_force_reader = csv.reader(pull_force_result)
            pull_force_list = list(pull_force_reader)
            # Run function
            report_status = self.update_pull_force_result(pull_force_list, format_sheet, format_row, format_column)
        pull_force_result.close()

        return report_status

    def update_pull_force_result(self, pull_force_list, format_sheet, format_row, format_column):
        data_req_list = ['1', '2', '3', '4', '5']
        report_status = "อัพเดทไปแล้ว"
        if str(format_sheet.cell(row=format_row, column=format_column).value) == "None":
            report_status = "โปรแกรมอัพเดท"
            for row in pull_force_list:
                if len(row) > 1 and row[0] in data_req_list:
                    format_sheet.cell(row=format_row, column=format_column).value = row[1]
                    format_row += 1
        return report_status

    # --------------------------------------- Copanarity Program ----------------------------------------
    def copanarity_window(self):
        self.hide_main_window()

        self.copanarity_gui = Toplevel()
        self.copanarity_gui.title('Copanarity Program')
        WIDTH = 900
        HEIGHT = 500
        screen_width = self.copanarity_gui.winfo_screenwidth()
        screen_height = self.copanarity_gui.winfo_screenheight()
        x = int((screen_width/2) - (WIDTH/2))
        y = int((screen_height/2) - (HEIGHT/1.8))
        self.copanarity_gui.resizable(0, 0)
        self.copanarity_gui.geometry(f'{WIDTH}x{HEIGHT}+{x}+{y}')
        self.copanarity_gui.protocol("WM_DELETE_WINDOW", self.copanarity_close_window)
        # Widget
        big_frame = Frame(self.copanarity_gui)
        topics_label = Label(big_frame, text='Copanarity Program', font=self.topic_font, fg='white', bg='blue')

        menu_frame = Frame(big_frame)
        year_label = Label(menu_frame, text='Year Report:', font=self.detail_font)
        self.year_cb = ttk.Combobox(menu_frame, font=self.detail_font, justify='center', width=self.combobox_width)
        month_label = Label(menu_frame, text='Month Report:', font=self.detail_font)
        self.month_cb = ttk.Combobox(menu_frame, font=self.detail_font, justify='center', width=self.combobox_width)
        # function add to combobox
        self.year_cb.bind('<Button-1>', self.copanarity_year_function)
        self.month_cb.bind('<Button-1>', self.copnarity_month_function)
        run_button = Button(big_frame, text='Run Program', command=self.copanarity_overview,
                            font=self.detail_font, width=18)

        treeview_frame = Frame(big_frame)
        headers = ['Filename', 'Item Code', 'FETL lot', 'Report status']
        self.copanarity_treeview = ttk.Treeview(treeview_frame, column=headers, show='headings',
                                                height=self.treeview_height)
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Arial', 16, 'bold'))
        style.configure("Treeview", font=('Arial', 12))
        vertical_scrollbar = ttk.Scrollbar(treeview_frame, orient="vertical", command=self.copanarity_treeview.yview)
        self.copanarity_treeview.configure(yscrollcommand=vertical_scrollbar.set)
        for header in headers:
            if header == 'Filename':
                column_width = 350
            else:
                column_width = 160
            self.copanarity_treeview.heading(header, text=header)
            self.copanarity_treeview.column(header, anchor='center', width=column_width, minwidth=0)
        self.copanarity_treeview.bind("<Double-1>", self.copanarity_link_tree)

        # Widget Position
        big_frame.pack()
        topics_label.grid(row=0, column=0, pady=5)
        # -- Menu Frame
        menu_frame.grid(row=1, column=0, pady=5)
        year_label.grid(row=0, column=0, padx=10, pady=5)
        self.year_cb.grid(row=0, column=1, padx=10, pady=5, ipady=2)
        month_label.grid(row=1, column=0, pady=5)
        self.month_cb.grid(row=1, column=1, pady=5, ipady=2)
        run_button.grid(row=2, column=0, pady=5)
        # --Treeview frame
        treeview_frame.grid(row=3, column=0, pady=5)
        self.copanarity_treeview.grid(row=0, column=0)
        vertical_scrollbar.grid(row=0, column=1, ipady=100)

    def copanarity_close_window(self):
        self.copanarity_gui.destroy()
        self.main_window()

    def copanarity_year_function(self, event):
        folder_list = []
        for folder_name in os.listdir(self.copanarity_machine_path):
            if os.path.isdir(os.path.join(self.copanarity_machine_path, folder_name)) and not str(folder_name).endswith('db'):
                folder_list.append(folder_name)
        self.year_cb['values'] = folder_list

    def copnarity_month_function(self, event):
        month_folder_list = []
        year = self.year_cb.get()

        if year == "":
            msb.showwarning(title='แจ้งเตือนไปยังผู้ใช้', message='กรุณากรอกช่อง Year report ให้เรียบร้อย')
        else:
            month_folder_path = os.path.join(self.copanarity_machine_path, year)
            folder_name_list = os.listdir(month_folder_path)

            for month_folder in folder_name_list:
                month_folder_location = os.path.join(month_folder_path, month_folder)

                if os.path.isdir(month_folder_location):
                    month_folder_list.append(month_folder)

        self.month_cb['values'] = month_folder_list

    def copanarity_link_tree(self, event):
        input_id = self.copanarity_treeview.selection()
        item_code = str(self.copanarity_treeview.item(input_id)["values"][1])
        fetl_lot = self.copanarity_treeview.item(input_id)["values"][2]
        report_status, copanarity_format_full_path = self.search_format_file(item_code, fetl_lot)
        try:
            webbrowser.open(copanarity_format_full_path)
        except Exception:
            msb.showwarning(title="Alarm to message", messsage=f'FETL Lot {fetl_lot} ไม่สามารถเปิดได้เนื่องจากไม่ได้สร้างไฟล์เอาไว้')

    def copanarity_overview(self):
        year = self.year_cb.get()
        month = self.month_cb.get()
        # Count number user not filled
        total_not_fill, status = self.copanarity_count_user_not_fill(year, month)

        if total_not_fill > 0:
            msb.showwarning(title='แจ้งเตือน', message='กรุณากรอกข้อมูลที่ ' + status + ' ก่อนรันโปรแกรม')
        else:
            self.clear_copanarity_treeview()
            self.update_status_to_copa_treeview(year, month)
            msb.showinfo(title='Information', message='Import complete')

    def copanarity_count_user_not_fill(self, year, month):
        total_not_fill = 0
        status = ""
        user_fill_dict = {'Year Report': year, 'Month Report': month}
        for fill_topic in user_fill_dict:
            if user_fill_dict[fill_topic] == "":
                total_not_fill += 1
                status = status + " " + fill_topic
        return total_not_fill, status

    def clear_copanarity_treeview(self):
        for member in self.copanarity_treeview.get_children():
            self.copanarity_treeview.delete(member)

    def update_status_to_copa_treeview(self, year, month):
        copanrity_result_full_path = os.path.join(os.path.join(self.copanarity_machine_path, year), month)
        for excel_file in os.listdir(copanrity_result_full_path):
            if str(excel_file).endswith('xls') and len(excel_file) == 34 and "LOT" in str(excel_file):
                # รูปแบบบันทึกใน excel: 31-08-20_13071373_LOT A8290025.xlsx
                item_code_first_index = excel_file.index('_') + 1
                fetl_lot_first_index = excel_file.index('T') + 2
                # ตัวแปรที่สำคัญในการรัน Function อื่นในโปรแกรม
                excel_file_path = os.path.join(copanrity_result_full_path, excel_file)
                item_code = excel_file[9:17]
                fetl_lot = excel_file[22:30]

                # Run function
                excel_wb, report_status, excel_ws = self.open_excel_file(excel_file_path)
                report_status, link = self.search_format_file(item_code, fetl_lot)
                
                # Check status for seperate operation
                if report_status == "Update" and "~$" not in link:
                    # Import data from machine result to format
                    format_wb, format_ws = self.open_copanarity_format(link)
                    row_format, copa1_column1, copa1_column2, copa2_column, copa3_column = self.search_copa_row_and_column(format_ws)
                    report_status = self.import_result_to_format(excel_ws, format_ws, row_format, copa1_column1, copa1_column2, copa2_column, copa3_column)
                    self.close_and_save_file(format_wb, link)

                    # Update to treeview
                    data = [excel_file, item_code, fetl_lot, report_status]
                    self.copanarity_treeview.insert('', 'end', value=data)

    def open_excel_file(self, excel_file_path):
        # มาแก้ต่อ
        excel_wb = xlrd.open_workbook(filename=excel_file_path)
        report_status, excel_ws = self.copanarity_column(excel_wb)
        return excel_wb, report_status, excel_ws

    def copanarity_column(self, excel_wb):
        # function นี้จะทำตรวจสอบว่ามี Copanarity ในเอกสารหรือไม่?
        if 'Sheet1' in excel_wb.sheet_names():
            # Update data when found copnarity
            report_status = "Update"
            excel_ws = excel_wb.sheet_by_name('Sheet1')
        else:
            excel_ws = ""
            report_status = "Not Copanarity Result"
        return report_status, excel_ws

    def search_format_file(self, item_code, fetl_lot):
        # ทำการค้นหาตำแหน่งไฟล์ของ format
        report_status = 'Not found'
        copanarity_format_full_path = '-'
        for year_folder in os.listdir(self.copanarity_machine_path):
            link = os.path.join(os.path.join(self.copanarity_format_path, year_folder), "01.Connector")
            if os.path.isdir(link):
                for item_code_folder in os.listdir(link):
                    if item_code == item_code_folder:
                        link2 = os.path.join(link, item_code_folder)
                        for fetl_lot_file in os.listdir(link2):
                            if fetl_lot in fetl_lot_file:
                                copanarity_format_full_path = os.path.join(link2, fetl_lot_file)
                                report_status = 'Update'
        return report_status, copanarity_format_full_path

    def open_copanarity_format(self, copanarity_format_full_path):
        # ทำการเปิดไฟล์ format ออกมา
        format_wb = xl.load_workbook(filename=copanarity_format_full_path)
        format_ws = format_wb[format_wb.sheetnames[0]]
        return format_wb, format_ws

    def search_copa_row_and_column(self, format_ws):
        # ทำการค้นหา แถวที่จะวางตำแหน่ง copanarity 1-3
        for row in range(1, format_ws.max_row+1):
            for column in range(1, format_ws.max_column+1):
                cell_value = str(format_ws.cell(row=row, column=column).value)
                if cell_value == "1":
                    row_format = row
                elif cell_value.startswith("Coplanarity") and cell_value.endswith("1"):
                    copa1_column1 = column
                    copa1_column2 = column + 1
                elif cell_value.startswith("Coplanarity") and cell_value.endswith("2"):
                    copa2_column = column
                elif cell_value.startswith("Coplanarity") and cell_value.endswith("3"):
                    copa3_column = column
        return row_format, copa1_column1, copa1_column2, copa2_column, copa3_column

    def import_result_to_format(self, excel_ws, format_ws, row_format, copa1_column1, copa1_column2, copa2_column, copa3_column):
        row = 1
        row_list = []
        report_status = "อัพเดทไปแล้ว"
        copa1_before = str(format_ws.cell(row=row_format, column=copa1_column1).value)
        copa1_after = str(format_ws.cell(row=row_format, column=copa1_column2).value)
        copa2_after = str(format_ws.cell(row=row_format, column=copa2_column).value)
        copa3_after = str(format_ws.cell(row=row_format, column=copa3_column).value)
        # ทำการตรวจสอบว่าใน format นั้น ไม่มีการอัพเดทข้อมูล
        if copa1_before == "None" and copa1_after == "None" and copa2_after == "None" and copa3_after == "None":
            # ทำการเพิ่มผล copanaity result ไปยัง list โดยมีเงื่อนไขว่าจะต้องมีคำว่า "C" อยู่ ==> C14
            report_status = "โปรแกรมอัพเดท"
            while row <= excel_ws.nrows-1:
                for column in range(excel_ws.ncols):
                    point_no = excel_ws.cell(rowx=0, colx=column).value
                    if str(point_no).startswith("C"):
                        copanrity_result = excel_ws.cell(rowx=row, colx=column).value
                        row_list.append(copanrity_result)
                # ทำการอัพเดทข้อมูลลงใน format
                different_copanarity1 = max(row_list)
                different_copanarity2 = different_copanarity1 + round(random.uniform(0.0001, 0.0030), 4)
                different_copanarity3 = different_copanarity2 + round(random.uniform(0.0001, 0.0030), 4)
                different_copanarity4 = different_copanarity3 + round(random.uniform(0.0001, 0.0030), 4)
                format_ws.cell(row=row_format, column=copa1_column1).value = different_copanarity1
                format_ws.cell(row=row_format, column=copa1_column2).value = different_copanarity2
                format_ws.cell(row=row_format, column=copa2_column).value = different_copanarity3
                format_ws.cell(row=row_format, column=copa3_column).value = different_copanarity4
                # ล้างข้อมูล list ทำการวนลูบในส่วนของ row_format = ข้อมูลที่จะวางถัดไป และ row ข้อมูลในแถวใน result
                row_list.clear()
                row_format += 1
                row += 1
        return report_status

    def close_and_save_file(self, format_wb, copanarity_format_full_path):
        # ทำการบันทึกผลข้อมูลและทำการปิด format
        format_wb.save(copanarity_format_full_path)
        format_wb.close()

    def close_excel_file(self, excel_wb):
        # ทำการปิด result format ที่เปิด
        excel_wb.release_resources()


# ทำการรัน class
app = SmtIqiProgram()
app.main_window()