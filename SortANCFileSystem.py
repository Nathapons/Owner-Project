import pandas as pd
from pyxlsb import open_workbook
from openpyxl import load_workbook
from os import path, listdir, mkdir
from shutil import copy2
from datetime import datetime

class SortAncFileSystem():
    def __init__(self):
        self.master_file_path = "\\\\10.17.78.169\\Main_DATA\\00.Record status ORT"
        self.barcode_error_path = '\\\\10.17.73.53\\ORT_Result\\37.OST\\แยก data แล้ว\\BARCODE ERROR'
        
    def program_overview(self):
        table_list = []

        for excel_file in listdir(self.master_file_path):
            if (excel_file.upper().startswith('ORT')):
                excel_file_location = path.join(self.master_file_path, excel_file)
                sheetnames = self.get_all_sheet(path=excel_file_location)
                table_list = self.get_table_list(sheetnames=sheetnames, table_list=table_list, filename=excel_file_location)

        self.get_filename(table_list)

    def get_all_sheet(self, path):
        sheetnames = []

        if path.endswith('xlsb'):
            wb = open_workbook(path)
            sheetnames = [name for name in wb.sheets if name not in ['ห้ามลบ', 'List Item', 'Operator']]
        else:
            wb = load_workbook(filename=path, read_only=True)
            for name in wb.sheetnames:
                if name not in ['ห้ามลบ', 'List Item', 'Operator']:
                    sheetnames.append(name)

        return sheetnames

    def get_table_list(self, filename, sheetnames, table_list):
        found_time = 0
        sheet_found = ''

        if filename.upper().endswith('XLSB'):
            engine ='pyxlsb'
        else:
            engine = None

        for sheetname in sheetnames:
            df = pd.read_excel(filename, sheet_name=sheetname, header=1, dtype=str, engine=engine)
            max_row = list(df.shape)[0]

            for row in range(max_row):
                barcode = str(df['S/N'][row])
                product_name = str(df['P/D name'][row]).strip()
                
                if "-" not in product_name and "Z" not in product_name:
                    first_path = product_name[0:3]
                    second_path = product_name[3:]
                    full_product_name = str(first_path) + "-" + str(second_path)
                elif "-" not in product_name:
                    first_path = product_name[0:4]
                    second_path = product_name[4:]
                    full_product_name = str(first_path) + "-" + str(second_path)
                else:
                    full_product_name = product_name
                
                lotno = str(df['lot no. for OST'][row])
                item_test = str(df['Item test'][row])

                row = [barcode, full_product_name, lotno, item_test]
                if 'nan' not in row:
                    table_list.append(row)

        return table_list

    def get_filename(self, table_list):
        yamaha_path = "\\\\10.17.73.53\\ORT_Result\\37.OST\\R2-40-131"
        yamaha_files = listdir(yamaha_path)

        for file in yamaha_files:
            if file.startswith('A') and file.upper().endswith('CSV'):
                barcode_no = file.split('_')[0]
                self.check_item_sort(table_list, barcode_no, file, yamaha_path)
                

    def check_item_sort(self, table_list, barcode_no, file, machine_path):
        category_list = [row for row in table_list if barcode_no in row]
        measure_list = list({row[3] for row in category_list})
        
        if len(category_list) > 0:
            if len(measure_list) == 1:
                last_index = len(category_list) - 1
                last_category_list = category_list[last_index]
                barcode = last_category_list[0]
                product = last_category_list[1]
                lotno = last_category_list[2]
                item_test = last_category_list[3]
                source = path.join(machine_path, file)

                self.create_folder(barcode, product, lotno, item_test, source)

    def create_folder(self, barcode, product, lotno, item_test, source):
        sorting_file_path = '\\\\10.17.73.53\\ORT_Result\\37.OST\\แยก data แล้ว'
        product_path = path.join(sorting_file_path, product)
        item_test_path = path.join(product_path, item_test)
        lotno_path = path.join(item_test_path, lotno)
        is_product_path = path.isdir(product_path)
        is_item_test_path = path.isdir(item_test_path)
        is_lotno_path = path.isdir(lotno_path)

        if is_product_path == False:
            mkdir(product_path)
        if is_item_test_path == False:
            mkdir(item_test_path)
        if is_lotno_path == False:
            mkdir(lotno_path)

        filename = "test_" + path.basename(source)
        destination = path.join(lotno_path, filename)
        is_file_exist = path.isfile(destination)
        
        if is_file_exist == False:
            copy2(source, destination)
            print(f'paste at {destination}')


app = SortAncFileSystem()
app.program_overview()